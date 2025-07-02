import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

def extract_sheet_name_from_filename(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    match = re.match(r"([A-Za-z]+)(\d{4})", base)
    if match:
        return f"{match.group(1)}_{match.group(2)}"
    return None

def generate_suggested_output_filename(filename, selected_modes):
    base = os.path.splitext(os.path.basename(filename))[0]
    match = re.match(r"([A-Za-z]+\d{4})", base)
    if match:
        base = match.group(1)

    suffix_parts = []
    if "Importer" in selected_modes:
        suffix_parts.append("im")
    if "Exporter" in selected_modes:
        suffix_parts.append("ex")

    suffix = "_" + "_".join(suffix_parts) + "_per_month"
    return base + suffix + ".xlsx"

def format_sheet(file_path, sheet_name, mode, unit):

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    blue_fill = PatternFill(start_color="D9E1F5", end_color="D9E1F5", fill_type="solid")

    ws.insert_cols(2)
    ws.insert_rows(1)

    ws.cell(row=1, column=1, value=f"Unit {unit.lower()}").fill = yellow_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=14)
    ws.cell(row=1, column=2, value="Months").fill = blue_fill
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value=mode).fill = blue_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=2, value="Total").fill = blue_fill
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for i, month in enumerate([f"{m:02d}" for m in range(1, 13)], start=3):
        ws.cell(row=2, column=i, value=month).fill = blue_fill
        ws.cell(row=2, column=i).alignment = Alignment(horizontal="center", vertical="center")

    # Fill row totals and apply number formatting
    for row in range(3, ws.max_row + 1):
        total = sum(ws.cell(row=row, column=col).value or 0 for col in range(3, 15))
        total_cell = ws.cell(row=row, column=2, value=round(total, 2))
        total_cell.number_format = '#,##0.00'

        for col in range(3, 15):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="Total").fill = blue_fill
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    total_sum = 0
    for row in range(3, total_row):
        val = ws.cell(row=row, column=2).value
        total_sum += val if isinstance(val, (int, float)) else 0

    ws.cell(row=total_row, column=2, value=round(total_sum, 2)).fill = blue_fill
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=2).number_format = '#,##0.00'

    # Add column totals (C to N)
    for col in range(3, 15):
        col_total = 0
        for row in range(3, total_row):
            val = ws.cell(row=row, column=col).value
            col_total += val if isinstance(val, (int, float)) else 0

        cell = ws.cell(row=total_row, column=col, value=round(col_total, 2))
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.number_format = '#,##0.00'

    wb.save(file_path)